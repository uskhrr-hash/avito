import unittest
from unittest.mock import MagicMock, patch

import pandas as pd

from avito.autoload import (
    DATA_START_ROW,
    filter_new_listings_workbook,
    filter_photo_updates_workbook,
    merge_autoload_feed_workbooks,
    save_workbook,
)
from avito.avito_api import (
    AvitoApiClient,
    AvitoApiConfig,
    fetch_avito_ids_by_ad_ids,
    update_item_price,
    update_stocks,
)
from avito.sync_listings import build_sync_items, sync_listings
from avito.stores import Store, StoresConfig
from openpyxl import Workbook


def _stores() -> StoresConfig:
    return StoresConfig(
        stores=(
            Store(prefix="md", label="md", fields={}),
            Store(prefix="pg", label="pg", fields={}),
        ),
        legacy_unprefixed_store="md",
    )


class TestSyncListings(unittest.TestCase):
    def test_build_sync_items_skips_without_avito_id(self):
        posting = pd.DataFrame(
            [
                {
                    "артикул": "100",
                    "номенклатура": "Tire A",
                    "recommended_price": 5000,
                    "количество": "8",
                },
                {
                    "артикул": "200",
                    "номенклатура": "Tire B",
                    "recommended_price": 6000,
                    "количество": "4",
                },
            ]
        )
        avito_ids = {"100": "111", "md_100": "111"}
        items = build_sync_items(posting, _stores(), avito_ids, max_listing_quantity=12)
        self.assertEqual(len(items), 2)
        self.assertEqual(items[0].listing_id, "md_100")
        self.assertEqual(items[0].avito_id, 111)
        self.assertEqual(items[0].price, 5000)
        self.assertEqual(items[0].quantity, 8)
        self.assertEqual(items[1].listing_id, "pg_100")
        self.assertEqual(items[1].avito_id, 111)

    def test_build_sync_items_caps_quantity(self):
        posting = pd.DataFrame(
            [
                {
                    "артикул": "100",
                    "номенклатура": "Tire",
                    "recommended_price": 5000,
                    "количество": "99",
                }
            ]
        )
        items = build_sync_items(
            posting,
            _stores(),
            {"md_100": "42"},
            max_listing_quantity=12,
        )
        self.assertEqual(items[0].quantity, 12)

    @patch("avito.sync_listings.update_stocks")
    @patch("avito.sync_listings.update_item_price")
    def test_sync_listings_calls_api(self, mock_price, mock_stocks):
        mock_stocks.return_value = [{"success": True}]
        from avito.sync_listings import SyncItem

        items = [
            SyncItem(
                listing_id="md_100",
                article="100",
                avito_id=111,
                price=5000,
                quantity=6,
            )
        ]
        client = MagicMock()
        stats = sync_listings(client, items, price_pause_sec=0)
        mock_price.assert_called_once_with(client, 111, 5000)
        mock_stocks.assert_called_once()
        self.assertEqual(stats.prices_updated, 1)
        self.assertEqual(stats.stocks_updated, 1)


class TestFilterNewListings(unittest.TestCase):
    def test_filter_keeps_only_rows_without_avito_id(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Шины"
        ws.cell(2, 1, "Уникальный идентификатор объявления")
        ws.cell(2, 2, "Номер объявления на Авито")
        ws.cell(2, 3, "Название объявления")
        ws.cell(DATA_START_ROW, 1, "md_100")
        ws.cell(DATA_START_ROW, 2, "111")
        ws.cell(DATA_START_ROW, 3, "Old tire")
        ws.cell(DATA_START_ROW + 1, 1, "md_200")
        ws.cell(DATA_START_ROW + 1, 3, "New tire")

        import tempfile
        from pathlib import Path

        with tempfile.TemporaryDirectory() as tmp:
            src = Path(tmp) / "all.xlsx"
            out = Path(tmp) / "new.xlsx"
            save_workbook(wb, src)
            kept, removed = filter_new_listings_workbook(
                src,
                out,
                avito_ids={"md_100": "111", "100": "111"},
            )
            self.assertEqual(kept, 1)
            self.assertEqual(removed, 1)
            wb2 = __import__("openpyxl").load_workbook(out)
            ws2 = wb2.active
            self.assertEqual(
                str(ws2.cell(DATA_START_ROW, 1).value),
                "md_200",
            )


class TestFilterPhotoUpdates(unittest.TestCase):
    def test_filter_keeps_existing_with_article_photos(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Шины"
        ws.cell(2, 1, "Уникальный идентификатор объявления")
        ws.cell(2, 2, "Номер объявления на Авито")
        ws.cell(2, 3, "Название объявления")
        ws.cell(2, 4, "Ссылки на фото")
        ws.cell(DATA_START_ROW, 1, "md_100")
        ws.cell(DATA_START_ROW, 2, "111")
        ws.cell(DATA_START_ROW, 3, "Old tire")
        ws.cell(
            DATA_START_ROW,
            4,
            "https://avito.shinaufa.ru/photos/md/md100-1.jpg",
        )
        ws.cell(DATA_START_ROW + 1, 1, "md_200")
        ws.cell(DATA_START_ROW + 1, 2, "222")
        ws.cell(DATA_START_ROW + 1, 3, "Model photo")
        ws.cell(
            DATA_START_ROW + 1,
            4,
            "https://avito.shinaufa.ru/photos/Yokohama Geolandar.jpg",
        )
        ws.cell(DATA_START_ROW + 2, 1, "md_300")
        ws.cell(DATA_START_ROW + 2, 3, "New tire")

        import tempfile
        from pathlib import Path

        with tempfile.TemporaryDirectory() as tmp:
            src = Path(tmp) / "all.xlsx"
            out = Path(tmp) / "photos.xlsx"
            save_workbook(wb, src)
            kept, removed = filter_photo_updates_workbook(
                src,
                out,
                avito_ids={"md_100": "111", "100": "111", "md_200": "222"},
            )
            self.assertEqual(kept, 1)
            self.assertGreater(removed, 0)
            wb2 = __import__("openpyxl").load_workbook(out)
            ws2 = wb2.active
            self.assertEqual(str(ws2.cell(DATA_START_ROW, 1).value), "md_100")

    def test_merge_feed_workbooks(self):
        import tempfile
        from pathlib import Path

        def _row(wb, lid, title):
            ws = wb.active
            ws.title = "Шины"
            ws.cell(2, 1, "Уникальный идентификатор объявления")
            ws.cell(2, 3, "Название объявления")
            row = ws.max_row + 1 if ws.max_row >= DATA_START_ROW else DATA_START_ROW
            if row < DATA_START_ROW:
                row = DATA_START_ROW
            ws.cell(row, 1, lid)
            ws.cell(row, 3, title)

        with tempfile.TemporaryDirectory() as tmp:
            p1 = Path(tmp) / "new.xlsx"
            p2 = Path(tmp) / "photo.xlsx"
            out = Path(tmp) / "merged.xlsx"
            wb1 = Workbook()
            _row(wb1, "md_1", "New")
            save_workbook(wb1, p1)
            wb2 = Workbook()
            _row(wb2, "md_2", "Photo update")
            save_workbook(wb2, p2)
            total = merge_autoload_feed_workbooks([p1, p2], out)
            self.assertEqual(total, 2)


class TestAvitoApiExtensions(unittest.TestCase):
    @patch("avito.avito_api.fetch_token")
    @patch("avito.avito_api.requests.request")
    def test_update_item_price(self, mock_req, mock_token):
        mock_token.return_value = __import__(
            "avito.avito_api", fromlist=["AvitoToken"]
        ).AvitoToken(access_token="t", expires_at=1e12)
        mock_req.return_value = MagicMock(status_code=200, content=b"{}", json=lambda: {})
        client = AvitoApiClient(AvitoApiConfig(client_id="a", client_secret="b"))
        update_item_price(client, 123, 5000)
        self.assertIn("/core/v1/items/123/update_price", mock_req.call_args.args[1])
        self.assertEqual(mock_req.call_args.kwargs["json"], {"price": 5000})

    @patch("avito.avito_api.fetch_token")
    @patch("avito.avito_api.requests.request")
    def test_fetch_avito_ids_by_ad_ids(self, mock_req, mock_token):
        mock_token.return_value = __import__(
            "avito.avito_api", fromlist=["AvitoToken"]
        ).AvitoToken(access_token="t", expires_at=1e12)
        mock_req.return_value = MagicMock(
            status_code=200,
            content=b'{"items":[{"ad_id":"md_1","avito_id":99}]}',
            json=lambda: {"items": [{"ad_id": "md_1", "avito_id": 99}]},
        )
        client = AvitoApiClient(AvitoApiConfig(client_id="a", client_secret="b"))
        out = fetch_avito_ids_by_ad_ids(client, ["md_1"])
        self.assertEqual(out, {"md_1": 99})


if __name__ == "__main__":
    unittest.main()
