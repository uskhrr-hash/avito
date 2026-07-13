import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import pandas as pd

from avito.autoload import (
    _apply_descriptions_to_sheet,
    _apply_listing_ids_to_sheet,
    _apply_quantities_to_sheet,
    _availability_headline,
    _payment_terms,
    _quantity_label,
    _remove_rows_without_photos,
    _sync_avito_ids_to_sheet,
    _sync_merge_ads_to_sheet,
    _sync_multi_names_to_sheet,
    _sync_photo_urls_to_sheet,
    _ensure_posting_rows_in_sheet,
    _apply_prices_to_sheet,
    _autoload_price,
    _format_description,
    avito_ids_for_posting,
    extract_avito_ids_from_xlsx,
    find_latest_avito_export,
    merge_avito_ids,
    resolve_autoload_base,
    normalize_article_id,
    posting_keep_sets,
    row_in_goods,
    save_workbook,
)
from openpyxl import Workbook
from avito.config import AutoloadSettings
from avito.stores import Store, StoresConfig


def _stores() -> StoresConfig:
    return StoresConfig(
        stores=(Store(prefix="md", label="md", fields={}),),
        legacy_unprefixed_store="md",
    )


class TestAutoload(unittest.TestCase):
    def test_posting_keep_listing_ids(self):
        df = pd.DataFrame([{"артикул": "103926", "номенклатура": "Tire"}])
        arts, titles, listing_ids = posting_keep_sets(df, _stores())
        self.assertIn("103926", arts)
        self.assertIn("md_103926", listing_ids)

    def test_autoload_price_rounds_to_tens(self):
        self.assertEqual(_autoload_price(5437), 5440)
        self.assertEqual(_autoload_price(5432), 5430)
        self.assertEqual(_autoload_price(4108), 4110)

    def test_apply_prices_syncs_from_posting(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(2, 1, "Название объявления")
        ws.cell(2, 2, "Цена")
        ws.cell(5, 1, "Kumho Ecowing ES31 195/65 R15 91H")
        ws.cell(5, 2, 4108)
        posting = pd.DataFrame(
            [{"номенклатура": "Kumho Ecowing ES31 195/65 R15 91H", "recommended_price": 4110}]
        )
        n = _apply_prices_to_sheet(ws, price_col=2, title_col=1, posting_df=posting)
        self.assertEqual(n, 1)
        self.assertEqual(ws.cell(5, 2).value, 4110)

    def test_row_in_goods_listing_id(self):
        self.assertTrue(
            row_in_goods(
                row_id="md_100",
                article="",
                title="",
                keep_articles=set(),
                keep_titles=set(),
                keep_listing_ids={"md_100"},
            )
        )

    def test_ensure_posting_rows_adds_missing(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(2, 1, "Уникальный идентификатор объявления")
        ws.cell(2, 2, "Название объявления")
        ws.cell(5, 1, "md_100")
        ws.cell(5, 2, "Tire A 205/55 R16")
        posting = pd.DataFrame(
            [
                {
                    "номенклатура": "Tire A 205/55 R16",
                    "артикул": "100",
                    "recommended_price": 4000,
                },
                {
                    "номенклатура": "Tire B 195/65 R15",
                    "артикул": "200",
                    "recommended_price": 5000,
                },
            ]
        )
        by_title = {"Tire A 205/55 R16": 5}
        by_id = {"md_100": 5, "100": 5}
        stats = {"appended": 0}
        proto = ["md_100", "Tire A", None, None, 4000]
        n = _ensure_posting_rows_in_sheet(
            ws,
            posting_df=posting,
            c_id=1,
            c_title=2,
            c_photos=3,
            c_avito=4,
            stores=_stores(),
            prototype_row=proto,
            by_title=by_title,
            by_id=by_id,
            next_row=6,
            stats=stats,
        )
        self.assertEqual(n, 1)
        self.assertEqual(ws.cell(6, 1).value, "md_200")
        self.assertEqual(ws.cell(6, 2).value, "Tire B 195/65 R15")
        self.assertFalse(str(ws.cell(6, 3).value or "").strip())

    def test_sync_avito_ids_clears_wrong_duplicate(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(2, 1, "Уникальный идентификатор объявления")
        ws.cell(2, 2, "Название объявления")
        ws.cell(2, 3, "Номер объявления на Авито")
        ws.cell(5, 1, "md_12044")
        ws.cell(5, 2, "Kumho Ecowing ES31 195/65 R15 91H")
        ws.cell(5, 3, "7956226221")
        ws.cell(6, 1, "md_149898")
        ws.cell(6, 2, "Cordiant Comfort 2 225/65 R17 106H")
        ws.cell(6, 3, "7956226221")
        posting = pd.DataFrame(
            [
                {
                    "номенклатура": "Kumho Ecowing ES31 195/65 R15 91H",
                    "артикул": "12044",
                    "recommended_price": 4110,
                },
                {
                    "номенклатура": "Cordiant Comfort 2 225/65 R17 106H",
                    "артикул": "149898",
                    "recommended_price": 5000,
                },
            ]
        )
        avito_ids = {"md_149898": "8011111111", "149898": "8011111111"}
        set_n, cleared = _sync_avito_ids_to_sheet(
            ws,
            avito_col=3,
            id_col=1,
            title_col=2,
            posting_df=posting,
            avito_ids=avito_ids,
        )
        self.assertEqual(str(ws.cell(6, 3).value), "8011111111")
        self.assertFalse(str(ws.cell(5, 3).value or "").strip())

    def test_remove_rows_without_photos(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(2, 2, "Название объявления")
        ws.cell(2, 3, "Ссылки на фото")
        ws.cell(5, 2, "With photo")
        ws.cell(5, 3, "yandex_disk://Авито/md1.jpg")
        ws.cell(6, 2, "No photo")
        ws.cell(6, 3, "")
        n = _remove_rows_without_photos(ws, photos_col=3, title_col=2)
        self.assertEqual(n, 1)
        self.assertEqual(ws.max_row, 5)
        self.assertEqual(ws.cell(5, 2).value, "With photo")

    def test_sync_photos_clears_copied_urls(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(2, 1, "Уникальный идентификатор объявления")
        ws.cell(2, 2, "Название объявления")
        ws.cell(2, 3, "Ссылки на фото")
        ws.cell(5, 1, "md_99999")
        ws.cell(5, 2, "No Photo Tire 205/55 R16")
        ws.cell(5, 3, "yandex_disk://Авито/md149898-1.jpg")
        posting = pd.DataFrame(
            [
                {
                    "номенклатура": "No Photo Tire 205/55 R16",
                    "артикул": "99999",
                    "recommended_price": 4000,
                }
            ]
        )
        cfg = AutoloadSettings(
            template_file=Path("t.xlsx"),
            working_file=Path("w.xlsx"),
            prefer_latest_avito_export=True,
            close_not_in_goods=False,
            sheet_name=None,
            image_mode="yandex_disk",
            yandex_disk_root="Авито",
            photos_public_base_url="",
            photo_layout="flat",
            photo_store_prefix_in_filename=True,
            image_count=0,
            image_ext="jpg",
            convert_photos_to_jpeg=True,
            jpeg_quality=92,
            compress_photos=True,
            jpeg_max_dimension=1920,
            compress_min_kb=400,
            model_photo_fallback=True,
            photo_article_first=True,
            manager_inbox_subdir="входящие",
            photos_local_dir=None,
            verify_photos_on_disk=False,
            model_descriptions_file=Path("m.xlsx"),
            missing_models_file="missing.xlsx",
            description_html="",
            store_pitch_html="<p>{company}</p>",
            llm_store_brief="",
            defaults={},
            skip_without_photos=True,
            include_all_goods_in_autoload=False,
            no_photos_file="no_photos.xlsx",
            avito_ids_file=Path("ids.csv"),
        )
        set_n, cleared = _sync_photo_urls_to_sheet(
            ws,
            photos_col=3,
            title_col=2,
            id_col=1,
            posting_df=posting,
            local_photos=None,
            cfg=cfg,
            stores=_stores(),
        )
        self.assertEqual(cleared, 1)
        self.assertFalse(str(ws.cell(5, 3).value or "").strip())

    def test_sync_photos_replaces_avito_hosted_with_article_photos(self):
        wb = Workbook()
        ws = wb.active
        avito_url = (
            "http://avito.ru/autoload/1/items-to-feed/images?"
            "imageSlug=keep-me"
        )
        ws.cell(2, 1, "Уникальный идентификатор объявления")
        ws.cell(2, 2, "Название объявления")
        ws.cell(2, 3, "Ссылки на фото")
        ws.cell(5, 1, "md_100")
        ws.cell(5, 2, "Tire 205/55 R16")
        ws.cell(5, 3, avito_url)
        posting = pd.DataFrame(
            [
                {
                    "номенклатура": "Tire 205/55 R16",
                    "артикул": "100",
                    "recommended_price": 4000,
                }
            ]
        )
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            (folder / "md100-1.jpg").write_bytes(b"x")
            cfg = AutoloadSettings(
                template_file=Path("t.xlsx"),
                working_file=Path("w.xlsx"),
                prefer_latest_avito_export=True,
                close_not_in_goods=False,
                sheet_name=None,
                image_mode="yandex_disk",
                yandex_disk_root="Авито",
            photos_public_base_url="",
                photo_layout="flat",
                photo_store_prefix_in_filename=True,
                image_count=0,
                image_ext="jpg",
                convert_photos_to_jpeg=True,
                jpeg_quality=92,
                compress_photos=True,
                jpeg_max_dimension=1920,
                compress_min_kb=400,
                model_photo_fallback=True,
                photo_article_first=True,
                manager_inbox_subdir="входящие",
                photos_local_dir=folder,
                verify_photos_on_disk=True,
                model_descriptions_file=Path("m.xlsx"),
                missing_models_file="missing.xlsx",
                description_html="",
                store_pitch_html="",
                llm_store_brief="",
                defaults={},
                skip_without_photos=True,
                include_all_goods_in_autoload=False,
                no_photos_file="no_photos.xlsx",
                avito_ids_file=Path("ids.csv"),
            )
            set_n, cleared = _sync_photo_urls_to_sheet(
                ws,
                photos_col=3,
                title_col=2,
                id_col=1,
                posting_df=posting,
                local_photos=folder,
                cfg=cfg,
                stores=_stores(),
            )
            self.assertEqual(set_n, 1)
            self.assertEqual(cleared, 0)
            self.assertIn("yandex_disk://", str(ws.cell(5, 3).value))
            self.assertNotEqual(str(ws.cell(5, 3).value), avito_url)

    def test_sync_photos_preserves_avito_hosted_for_model_only(self):
        wb = Workbook()
        ws = wb.active
        avito_url = (
            "http://avito.ru/autoload/1/items-to-feed/images?"
            "imageSlug=model-on-avito"
        )
        ws.cell(2, 1, "Уникальный идентификатор объявления")
        ws.cell(2, 2, "Название объявления")
        ws.cell(2, 3, "Ссылки на фото")
        ws.cell(5, 1, "md_99999")
        ws.cell(5, 2, "Formula Energy 205/55 R16 91H")
        ws.cell(5, 3, avito_url)
        posting = pd.DataFrame(
            [
                {
                    "номенклатура": "Formula Energy 205/55 R16 91H",
                    "артикул": "99999",
                    "recommended_price": 4000,
                }
            ]
        )
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            (folder / "Formula Energy.jpg").write_bytes(b"x")
            cfg = AutoloadSettings(
                template_file=Path("t.xlsx"),
                working_file=Path("w.xlsx"),
                prefer_latest_avito_export=True,
                close_not_in_goods=False,
                sheet_name=None,
                image_mode="yandex_disk",
                yandex_disk_root="Авито",
                photos_public_base_url="",
                photo_layout="flat",
                photo_store_prefix_in_filename=True,
                image_count=0,
                image_ext="jpg",
                convert_photos_to_jpeg=True,
                jpeg_quality=92,
                compress_photos=True,
                jpeg_max_dimension=1920,
                compress_min_kb=400,
                model_photo_fallback=True,
                photo_article_first=True,
                manager_inbox_subdir="входящие",
                photos_local_dir=folder,
                verify_photos_on_disk=True,
                model_descriptions_file=Path("m.xlsx"),
                missing_models_file="missing.xlsx",
                description_html="",
                store_pitch_html="",
                llm_store_brief="",
                defaults={},
                skip_without_photos=True,
                include_all_goods_in_autoload=False,
                no_photos_file="no_photos.xlsx",
                avito_ids_file=Path("ids.csv"),
            )
            set_n, cleared = _sync_photo_urls_to_sheet(
                ws,
                photos_col=3,
                title_col=2,
                id_col=1,
                posting_df=posting,
                local_photos=folder,
                cfg=cfg,
                stores=_stores(),
            )
            self.assertEqual(set_n, 0)
            self.assertEqual(cleared, 0)
            self.assertEqual(str(ws.cell(5, 3).value), avito_url)

    def test_apply_listing_ids_replaces_avito_numeric_id(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(2, 1, "Уникальный идентификатор объявления")
        ws.cell(2, 2, "Название объявления")
        ws.cell(5, 1, "8020124332")
        ws.cell(5, 2, "Kumho Ecowing ES31 195/65 R15 91H")
        ws.cell(6, 1, "8020124332")
        ws.cell(6, 2, "Ikon Tyres Autograph Eco 3 205/55 R16 94H")
        posting = pd.DataFrame(
            [
                {
                    "номенклатура": "Kumho Ecowing ES31 195/65 R15 91H",
                    "артикул": "12044",
                    "recommended_price": 4110,
                },
                {
                    "номенклатура": "Ikon Tyres Autograph Eco 3 205/55 R16 94H",
                    "артикул": "149898",
                    "recommended_price": 5000,
                },
            ]
        )
        n = _apply_listing_ids_to_sheet(
            ws,
            id_col=1,
            title_col=2,
            posting_df=posting,
            stores=_stores(),
        )
        self.assertEqual(n, 2)
        self.assertEqual(ws.cell(5, 1).value, "md_12044")
        self.assertEqual(ws.cell(6, 1).value, "md_149898")

    def test_resolve_base_prefers_latest_export(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "input"
            input_dir.mkdir()
            old = input_dir / "432801655_2026-06-01T00_00_00Z.xlsx"
            new = input_dir / "432801655_2026-06-10T09_50_00Z.xlsx"
            old.write_bytes(b"old")
            new.write_bytes(b"newer")
            working = root / "input" / "autoload_working.xlsx"
            working.write_bytes(b"working")
            cfg = AutoloadSettings(
                template_file=Path("input/fallback.xlsx"),
                working_file=Path("input/autoload_working.xlsx"),
                prefer_latest_avito_export=True,
                close_not_in_goods=False,
                sheet_name=None,
                image_mode="yandex_disk",
                yandex_disk_root="Авито",
            photos_public_base_url="",
                photo_layout="flat",
                photo_store_prefix_in_filename=True,
                image_count=0,
                image_ext="jpg",
                convert_photos_to_jpeg=True,
            jpeg_quality=92,
            compress_photos=True,
            jpeg_max_dimension=1920,
            compress_min_kb=400,
            model_photo_fallback=True,
            photo_article_first=True,
            manager_inbox_subdir="входящие",
            photos_local_dir=None,
                verify_photos_on_disk=False,
                model_descriptions_file=Path("m.xlsx"),
                missing_models_file="missing.xlsx",
                description_html="<p>{nomenclature}</p>",
                store_pitch_html="",
                llm_store_brief="",
                defaults={},
                skip_without_photos=True,
                include_all_goods_in_autoload=False,
                no_photos_file="no_photos.xlsx",
                avito_ids_file=Path("ids.csv"),
            )
            path, label = resolve_autoload_base(cfg, root=root)
            self.assertEqual(path, new)
            self.assertIn("последняя выгрузка", label)

            path2, _ = resolve_autoload_base(cfg, root=root, use_working=True)
            self.assertEqual(path2, working)

    def test_extract_avito_ids_from_xlsx(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Шины, диски и колёса"
        ws.cell(2, 1, "Уникальный идентификатор объявления")
        ws.cell(2, 3, "Номер объявления на Авито")
        ws.cell(5, 1, "md_12044")
        ws.cell(5, 3, "7956581577")
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "export.xlsx"
            wb.save(path)
            got = extract_avito_ids_from_xlsx(path, _stores())
            self.assertEqual(got["md_12044"], "7956581577")
            self.assertEqual(got["12044"], "7956581577")

    def test_merge_avito_ids_csv_overrides(self):
        merged = merge_avito_ids(
            {"12044": "111"},
            {"12044": "222"},
            stores=_stores(),
        )
        self.assertEqual(merged["12044"], "222")

    def test_avito_ids_by_title_from_export(self):
        posting = pd.DataFrame(
            [
                {
                    "номенклатура": "Kumho Ecowing ES31 195/65 R15 91H",
                    "артикул": "12044",
                    "recommended_price": 4110,
                }
            ]
        )
        got = avito_ids_for_posting(
            posting,
            _stores(),
            titles_from_xlsx={
                "Kumho Ecowing ES31 195/65 R15 91H": "7956581577",
            },
        )
        self.assertEqual(got["12044"], "7956581577")
        self.assertEqual(got["md_12044"], "7956581577")

    def test_format_description_includes_marketing_block(self):
        tpl = (
            '<p><strong>{availability_headline}</strong></p>'
            '<p>Новые шины &quot;{nomenclature}&quot;🛞🛞🛞</p>'
            "<p>{model_description}</p>"
        )
        out = _format_description(
            tpl,
            nomenclature="Kumho Ecowing ES31 195/65 R15 91H",
            article="12044",
            price=4110,
            quantity="4",
            model_description="<p>Описание модели</p>",
            store_defaults={"phone": "79273181543"},
            ushk_in_stock=True,
        )
        self.assertIn("Шины в наличии!", out)
        self.assertIn("Kumho Ecowing ES31 195/65 R15 91H", out)
        self.assertIn("🛞", out)
        self.assertIn("Описание модели", out)

    def test_format_description_order_when_no_ushk(self):
        tpl = "<p><strong>{availability_headline}</strong></p>"
        out = _format_description(
            tpl,
            nomenclature="Tire",
            article="1",
            price=5000,
            quantity="4",
            model_description="",
            store_defaults={},
            ushk_in_stock=False,
        )
        self.assertIn("Шины под заказ 1-2 дня", out)

    def test_availability_headline(self):
        self.assertEqual(_availability_headline(True), "Шины в наличии!")
        self.assertEqual(_availability_headline(False), "Шины под заказ 1-2 дня")

    def test_payment_terms(self):
        self.assertEqual(_payment_terms(True), "Цена за наличный расчет")
        self.assertEqual(_payment_terms(False), "Любая форма оплаты, НДС")

    def test_format_description_payment_terms(self):
        tpl = "<p><strong>{payment_terms}</strong></p>"
        out = _format_description(
            tpl,
            nomenclature="Tire",
            article="1",
            price=5000,
            quantity="4",
            model_description="",
            store_defaults={},
            sam_mb_cash_price=True,
        )
        self.assertIn("Цена за наличный расчет", out)

    def test_sync_merge_ads_enables_multiitem(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(2, 1, "Уникальный идентификатор объявления")
        ws.cell(2, 2, "Соединять это объявление с другими объявлениями")
        ws.cell(5, 1, "md_100")
        ws.cell(5, 2, "Нет")
        headers = {"Уникальный идентификатор объявления": 1, "Соединять это объявление с другими объявлениями": 2}
        n = _sync_merge_ads_to_sheet(ws, headers)
        self.assertEqual(n, 1)
        self.assertEqual(ws.cell(5, 2).value, "Да")

    def test_sync_multi_names_groups_by_size(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(2, 1, "Название объявления")
        ws.cell(2, 2, "Название мультиобъявления")
        ws.cell(5, 1, "Kumho Ecowing ES31 195/65 R15 91H")
        ws.cell(6, 1, "Nokian Nordman 7 195/65 R15 95H")
        ws.cell(7, 1, "Yokohama Geolandar G015 245/70 R16 111H")
        headers = {"Название объявления": 1, "Название мультиобъявления": 2}
        changed, groups = _sync_multi_names_to_sheet(ws, headers, title_col=1)
        self.assertEqual(changed, 3)
        self.assertEqual(groups, 2)
        self.assertEqual(ws.cell(5, 2).value, "19565R15")
        self.assertEqual(ws.cell(6, 2).value, "19565R15")
        self.assertEqual(ws.cell(7, 2).value, "24570R16")

    def test_apply_descriptions_updates_all_rows(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(2, 1, "Уникальный идентификатор объявления")
        ws.cell(2, 2, "Название объявления")
        ws.cell(2, 3, "Цена")
        ws.cell(2, 4, "Описание объявления")
        ws.cell(5, 1, "md_12044")
        ws.cell(5, 2, "Kumho Ecowing ES31 195/65 R15 91H")
        ws.cell(5, 3, 4110)
        ws.cell(5, 4, "старое описание")
        posting = pd.DataFrame(
            [
                {
                    "номенклатура": "Kumho Ecowing ES31 195/65 R15 91H",
                    "артикул": "12044",
                    "recommended_price": 4110,
                    "количество": 4,
                    "ушк_в_наличии": True,
                }
            ]
        )
        tpl = '<p><strong>{availability_headline}</strong></p><p>{nomenclature}</p>'
        from avito.config import AutoloadSettings
        from pathlib import Path

        cfg = AutoloadSettings(
            template_file=Path("t.xlsx"),
            working_file=Path("w.xlsx"),
            prefer_latest_avito_export=True,
            close_not_in_goods=False,
            sheet_name=None,
            image_mode="yandex_disk",
            yandex_disk_root="Авито",
            photos_public_base_url="",
            photo_layout="flat",
            photo_store_prefix_in_filename=True,
            image_count=0,
            image_ext="jpg",
            convert_photos_to_jpeg=True,
            jpeg_quality=92,
            compress_photos=True,
            jpeg_max_dimension=1920,
            compress_min_kb=400,
            model_photo_fallback=True,
            photo_article_first=True,
            manager_inbox_subdir="входящие",
            photos_local_dir=None,
            verify_photos_on_disk=False,
            model_descriptions_file=Path("m.xlsx"),
            missing_models_file="missing.xlsx",
            description_html=tpl,
            store_pitch_html="",
            llm_store_brief="",
            defaults={"phone": "79273181543"},
            skip_without_photos=True,
            include_all_goods_in_autoload=False,
            no_photos_file="no_photos.xlsx",
            avito_ids_file=Path("ids.csv"),
        )
        n = _apply_descriptions_to_sheet(
            ws,
            desc_col=4,
            title_col=2,
            price_col=3,
            id_col=1,
            qty_col=None,
            posting_df=posting,
            cfg=cfg,
            stores=_stores(),
            model_descriptions={},
        )
        self.assertEqual(n, 1)
        self.assertIn("Шины в наличии!", str(ws.cell(5, 4).value))

    def test_quantity_label_caps_at_max(self):
        self.assertEqual(_quantity_label("20", max_quantity=12), "12")
        self.assertEqual(_quantity_label("4", max_quantity=12), "4")
        self.assertEqual(_quantity_label("", max_quantity=12), "1")

    def test_apply_quantities_always_one_for_price(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(2, 1, "Название объявления")
        ws.cell(2, 2, "Количество")
        ws.cell(5, 1, "Kumho Ecowing ES31 195/65 R15 91H")
        ws.cell(5, 2, "6")
        posting = pd.DataFrame(
            [
                {
                    "номенклатура": "Kumho Ecowing ES31 195/65 R15 91H",
                    "артикул": "12044",
                    "recommended_price": 4110,
                    "количество": 20,
                }
            ]
        )
        n = _apply_quantities_to_sheet(
            ws,
            qty_col=2,
            title_col=1,
            posting_df=posting,
            max_quantity=12,
        )
        self.assertEqual(n, 1)
        self.assertEqual(str(ws.cell(5, 2).value), "1")

    def test_save_workbook_fallback_when_locked(self):
        with tempfile.TemporaryDirectory() as tmp:
            target = Path(tmp) / "autoload.xlsx"
            wb = Workbook()
            real_save = wb.save
            calls: list[Path] = []

            def fake_save(path):
                calls.append(Path(path))
                if len(calls) == 1:
                    raise PermissionError(13, "Permission denied")
                real_save(path)

            with patch.object(wb, "save", side_effect=fake_save):
                saved = save_workbook(wb, target)
            self.assertNotEqual(saved, target)
            self.assertTrue(saved.is_file())


if __name__ == "__main__":
    unittest.main()
