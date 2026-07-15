#!/usr/bin/env python3
"""Копирует autoload xlsx в каталог фида на VPS и запускает upload через API."""
from __future__ import annotations

import argparse
import json
import logging
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from avito.autoload import (
    DATA_START_ROW,
    _col,
    _extract_avito_export_maps,
    _find_data_sheet,
    _header_map,
    avito_ids_for_posting,
    find_latest_avito_export,
    load_avito_ids,
    load_posting,
    merge_autoload_feed_workbooks,
    normalize_article_id,
    save_avito_ids_csv,
)
from avito.avito_api import (
    AvitoApiClient,
    DEFAULT_AUTOLOAD_SCHEDULE,
    get_autoload_profile,
    get_last_successful_upload,
    load_avito_api_config,
    trigger_autoload_upload,
    update_autoload_profile,
)
from avito.config import load_config, load_merged_yaml
from avito.db import load_secrets
from avito.sync_listings import (
    build_sync_items,
    refresh_avito_ids_from_api,
    sync_listings,
)

LOG = logging.getLogger("publish_avito_feed")


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Публикация фида Avito (файл + API upload)")
    p.add_argument("-c", "--config", type=Path, default=ROOT / "config.yaml")
    p.add_argument(
        "--source",
        type=Path,
        default=None,
        help="autoload xlsx (по умолчанию input/autoload_new.xlsx — только новые)",
    )
    p.add_argument("--dry-run", action="store_true")
    p.add_argument("--no-upload", action="store_true")
    p.add_argument("--no-profile", action="store_true")
    p.add_argument("--no-sync", action="store_true", help="Не обновлять цены/остатки через API")
    return p.parse_args()


def _load_publish_cfg(config_path: Path) -> dict:
    raw = load_merged_yaml(config_path)
    return dict(raw.get("avito_publish") or {})


def _resolve_report_email(profile: dict, pub: dict) -> str:
    return str(profile.get("report_email") or pub.get("report_email") or "").strip()


def _resolve_schedule(profile: dict, pub: dict) -> list[dict]:
    sched = profile.get("schedule") or pub.get("schedule")
    if sched:
        return list(sched)
    return list(DEFAULT_AUTOLOAD_SCHEDULE)


def _resolve_feed_path(config_parent: Path, path: Path | None) -> Path | None:
    if path is None:
        return None
    if not path.is_absolute():
        path = config_parent / path
    return path


def _collect_publish_sources(app, config_path: Path) -> tuple[list[Path], int, int]:
    """Новые объявления + обновления фото (разные файлы, без пересечения по AvitoId)."""
    parent = config_path.parent
    new_path = _resolve_feed_path(parent, app.autoload.new_listings_feed)
    photo_path = _resolve_feed_path(parent, app.autoload.photo_updates_feed)
    new_count = _count_data_rows(new_path) if new_path and new_path.is_file() else 0
    photo_count = _count_data_rows(photo_path) if photo_path and photo_path.is_file() else 0
    sources: list[Path] = []
    if new_count > 0 and new_path:
        sources.append(new_path)
    if photo_count > 0 and photo_path:
        sources.append(photo_path)
    return sources, new_count, photo_count


def _count_data_rows(path: Path) -> int:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = _find_data_sheet(wb)
        ws = wb[sheet_name]
        headers = _header_map(ws)
        c_id = _col(headers, "Уникальный идентификатор объявления")
        c_title = _col(headers, "Название объявления")
        count = 0
        for row_idx in range(DATA_START_ROW, ws.max_row + 1):
            listing_id = normalize_article_id(ws.cell(row_idx, c_id).value if c_id else None)
            title = str(ws.cell(row_idx, c_title).value or "").strip() if c_title else ""
            if listing_id or title:
                count += 1
        return count
    finally:
        wb.close()


def _listing_ids_from_feed(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = _find_data_sheet(wb)
        ws = wb[sheet_name]
        headers = _header_map(ws)
        c_id = _col(headers, "Уникальный идентификатор объявления")
        if not c_id:
            return []
        out: list[str] = []
        for row_idx in range(DATA_START_ROW, ws.max_row + 1):
            listing_id = normalize_article_id(ws.cell(row_idx, c_id).value)
            if listing_id and "_" in listing_id:
                out.append(listing_id)
        return out
    finally:
        wb.close()


def _run_api_sync(app, config_path: Path, *, dry_run: bool) -> int:
    sync_cfg = app.avito_sync
    if not sync_cfg.enabled:
        LOG.info("avito_sync.enabled=false — пропуск API-синхронизации")
        return 0

    posting_files = sorted(
        (ROOT / "output").glob("posting_*.xlsx"),
        key=lambda p: p.stat().st_mtime,
    )
    if not posting_files:
        LOG.warning("Нет posting_*.xlsx — пропуск API-синхронизации")
        return 0

    posting_df = load_posting(posting_files[-1])
    avito_ids_path = ROOT / app.autoload.avito_ids_file
    ids_from_csv = load_avito_ids(avito_ids_path, app.stores) if avito_ids_path.exists() else {}
    export = find_latest_avito_export(ROOT / "input")
    ids_from_xlsx: dict[str, str] = {}
    titles_from_xlsx: dict[str, str] = {}
    if export:
        ids_from_xlsx, titles_from_xlsx = _extract_avito_export_maps(export)
    avito_ids = avito_ids_for_posting(
        posting_df,
        app.stores,
        ids_from_xlsx=ids_from_xlsx,
        titles_from_xlsx=titles_from_xlsx,
        ids_from_csv=ids_from_csv,
    )
    items = build_sync_items(
        posting_df,
        app.stores,
        avito_ids,
        max_listing_quantity=app.autoload.max_listing_quantity,
    )
    if not items:
        LOG.info("Нет объявлений для API-синхронизации")
        return 0

    LOG.info("API: обновление цены/остатка для %s уже размещённых объявлений", len(items))
    secrets_path = app.stock_sources.secrets_file
    if not secrets_path.is_absolute():
        secrets_path = config_path.parent / secrets_path
    client = AvitoApiClient(load_avito_api_config(load_secrets(secrets_path)))
    stats = sync_listings(
        client,
        items,
        dry_run=dry_run or sync_cfg.dry_run,
        stock_batch_size=sync_cfg.stock_batch_size,
        price_pause_sec=sync_cfg.price_pause_sec,
    )
    LOG.info(
        "API sync: цены ok=%s fail=%s | остатки ok=%s fail=%s",
        stats.prices_updated,
        stats.prices_failed,
        stats.stocks_updated,
        stats.stocks_failed,
    )
    for err in stats.errors[:10]:
        LOG.warning("%s", err)
    if stats.prices_failed or stats.stocks_failed:
        LOG.warning(
            "API sync: частичные ошибки (цены fail=%s, остатки fail=%s) — пайплайн не останавливаем",
            stats.prices_failed,
            stats.stocks_failed,
        )
    return 0


def main() -> int:
    args = parse_args()
    logging.basicConfig(level=logging.INFO, format="%(message)s")

    pub = _load_publish_cfg(args.config)
    if not pub.get("enabled", False) and not args.dry_run:
        LOG.error(
            "avito_publish.enabled=false (проверьте config.yaml и config.local.yaml)"
        )
        return 1

    feed_url = str(pub.get("feed_public_url", "")).strip()
    feed_name = str(pub.get("feed_name", "shinaufa")).strip()
    feed_dir = Path(str(pub.get("feed_local_dir", "/var/www/avito-feed/feeds")))
    if not feed_url:
        LOG.error("Задайте avito_publish.feed_public_url")
        return 1

    app = load_config(args.config)
    new_feed_for_ids: Path | None = None
    if args.source is not None:
        source = args.source
        if not source.is_absolute():
            source = args.config.parent / source
        if not source.is_file():
            LOG.error("Нет файла фида: %s", source)
            return 1
        publish_sources = [source]
        new_count = _count_data_rows(source)
        photo_count = 0
        new_feed_for_ids = source
    else:
        publish_sources, new_count, photo_count = _collect_publish_sources(
            app, args.config
        )
        new_path = _resolve_feed_path(
            args.config.parent, app.autoload.new_listings_feed
        )
        if new_path and new_path.is_file():
            new_feed_for_ids = new_path
        if not publish_sources:
            LOG.error(
                "Нет файлов для публикации (сначала build_autoload.py: "
                "autoload_new.xlsx / autoload_photo_updates.xlsx)"
            )
            return 1

    if args.dry_run:
        LOG.info(
            "dry-run: новых=%s, обновление фото=%s → %s",
            new_count,
            photo_count,
            feed_url,
        )
        return 0

    sync_rc = 0
    if not args.no_sync:
        sync_rc = _run_api_sync(app, args.config, dry_run=False)

    if new_count == 0 and photo_count == 0:
        LOG.info(
            "Новых объявлений и обновлений фото нет — upload пропущен (дубли не создаём)"
        )
        return sync_rc

    target = feed_dir / "autoload.xlsx"
    feed_dir.mkdir(parents=True, exist_ok=True)
    row_count = merge_autoload_feed_workbooks(publish_sources, target)
    LOG.info(
        "Фид: новых=%s, обновление фото=%s, всего=%s строк → %s",
        new_count,
        photo_count,
        row_count,
        target,
    )
    LOG.info("Скопировано (%s байт)", target.stat().st_size)

    secrets_path = app.stock_sources.secrets_file
    if not secrets_path.is_absolute():
        secrets_path = args.config.parent / secrets_path
    client = AvitoApiClient(load_avito_api_config(load_secrets(secrets_path)))

    profile = get_autoload_profile(client)
    report_email = _resolve_report_email(profile, pub)
    schedule = _resolve_schedule(profile, pub)
    feeds = profile.get("feeds_data") or []
    need_profile = not feeds or not any(
        str(f.get("feed_url", "")).strip() == feed_url for f in feeds if isinstance(f, dict)
    )

    if need_profile and not args.no_profile and pub.get("auto_set_profile", True):
        if not report_email:
            LOG.error(
                "Нет report_email — добавьте в config.local.yaml:\n"
                "  avito_publish:\n"
                "    report_email: ваш@email.ru"
            )
            LOG.warning("Пропускаем обновление профиля (фид на сервере уже скопирован)")
        else:
            LOG.info("Обновляем профиль автозагрузки → %s", feed_url)
            try:
                update_autoload_profile(
                    client,
                    feed_name=feed_name,
                    feed_url=feed_url,
                    report_email=report_email,
                    schedule=schedule,
                )
            except (RuntimeError, ValueError) as exc:
                LOG.error("Профиль не обновлён: %s", exc)
                LOG.warning(
                    "Фид уже на %s — можно задать URL вручную в ЛК Авито "
                    "или исправить report_email/schedule",
                    feed_url,
                )
    elif need_profile:
        LOG.warning("feeds_data пустой — включите auto_set_profile или настройте URL в ЛК")

    if not args.no_upload and pub.get("auto_upload", True):
        LOG.info("Запуск upload (новые + обновление фото)…")
        try:
            trigger_autoload_upload(client)
            LOG.info("upload принят Avito")
        except RuntimeError as exc:
            if "429" in str(exc) or "час" in str(exc).lower():
                LOG.warning("%s (лимит 1 раз/час — нормально)", exc)
            else:
                raise
        try:
            last = get_last_successful_upload(client)
            if last:
                LOG.info(
                    "Последняя успешная загрузка:\n%s",
                    json.dumps(last, ensure_ascii=False, indent=2)[:2000],
                )
        except Exception as exc:
            LOG.warning("last_successful upload: %s", exc)

        if app.avito_sync.refresh_ids_after_publish and new_feed_for_ids:
            ad_ids = _listing_ids_from_feed(new_feed_for_ids)
            if ad_ids:
                avito_ids_path = ROOT / app.autoload.avito_ids_file
                existing = (
                    load_avito_ids(avito_ids_path, app.stores)
                    if avito_ids_path.exists()
                    else {}
                )
                merged = refresh_avito_ids_from_api(
                    client,
                    ad_ids,
                    existing=existing,
                    stores=app.stores,
                )
                new_count = sum(1 for k in merged if k not in existing)
                if merged:
                    save_avito_ids_csv(avito_ids_path, merged)
                    LOG.info(
                        "avito_ids.csv: %s записей (новых с API: %s)",
                        len({k for k in merged if "_" not in k}),
                        new_count,
                    )

    return sync_rc


if __name__ == "__main__":
    raise SystemExit(main())
