"""Заполнение шаблона автозагрузки Avito (Excel)."""
from __future__ import annotations

import copy
import logging
import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from avito.config import AutoloadSettings
from avito.photos import (
    PhotoNamingSettings,
    build_store_photo_urls,
    is_avito_hosted_photo_urls,
    normalize_yandex_photo_urls,
    photo_urls_look_like_article,
    resolve_listing_photo_sets,
)
from avito.yandex_disk_api import YandexDiskDownloadUrls, load_yandex_oauth_token

LOG = logging.getLogger(__name__)
from avito.stores import Store, StoresConfig, merge_defaults
from avito.model_descriptions import lookup_model_description
from avito.pricing import round_price_to_tens
from avito.title_parse import parse_title_fields, build_multi_name_from_title

# Строки шаблона (1-based): 1 категория, 2 заголовки, 3 обязательность, 4 подсказки, 5+ данные
DATA_START_ROW = 5
HEADER_ROW = 2


def save_workbook(wb, path: Path) -> Path:
    """
    Сохранить xlsx. Если файл занят (часто открыт в Excel) — резервное имя.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(path)
        return path
    except PermissionError:
        alt = path.with_name(f"{path.stem}_{datetime.now():%H%M%S}{path.suffix}")
        LOG.warning(
            "Не удалось записать %s — файл занят (закройте в Excel). "
            "Сохраняю как %s",
            path,
            alt.name,
        )
        wb.save(alt)
        return alt


def _find_data_sheet(wb) -> str:
    for name in wb.sheetnames:
        if name.startswith("Спр"):
            continue
        if name == "Инструкция":
            continue
        if "Легковые" in name or "Шины" in name:
            return name
    for name in wb.sheetnames:
        if not name.startswith("Спр") and name != "Инструкция":
            return name
    raise ValueError("Не найден лист шаблона в файле автозагрузки")


def _header_map(ws) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(HEADER_ROW, col).value
        if val is None:
            continue
        key = str(val).strip()
        if key:
            mapping[key] = col
    return mapping


def _col(headers: dict[str, int], name: str) -> int | None:
    if name in headers:
        return headers[name]
    low = name.lower()
    for k, v in headers.items():
        if k.lower() == low:
            return v
    return None


def normalize_article_id(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return ""
    if s.endswith(".0"):
        try:
            return str(int(float(s)))
        except ValueError:
            pass
    return s


def posting_keep_sets(
    posting_df: pd.DataFrame,
    stores: StoresConfig,
) -> tuple[set[str], set[str], set[str]]:
    """Артикулы, номенклатура и Id объявления (md_артикул) из posting."""
    articles: set[str] = set()
    titles: set[str] = set()
    listing_ids: set[str] = set()
    for _, post in posting_df.iterrows():
        nom = str(post.get("номенклатура", "")).strip()
        if nom and nom.lower() != "nan":
            titles.add(nom)
        article = normalize_article_id(post.get("артикул", ""))
        if article:
            articles.add(article)
            listing_ids.add(article)
            for store in stores.stores:
                listing_ids.add(store.listing_id(article))
    return articles, titles, listing_ids


def row_in_goods(
    *,
    row_id: str,
    article: str,
    title: str,
    keep_articles: set[str],
    keep_titles: set[str],
    keep_listing_ids: set[str],
) -> bool:
    if row_id and row_id in keep_listing_ids:
        return True
    if article and article in keep_articles:
        return True
    if title and title in keep_titles:
        return True
    return False


def resolve_photos_folder(cfg: AutoloadSettings, project_root: Path) -> Path | None:
    if not cfg.verify_photos_on_disk or not cfg.photos_local_dir:
        return None
    folder = cfg.photos_local_dir
    if not folder.is_absolute():
        folder = project_root / folder
    return folder if folder.is_dir() else None


def _photo_cfg(cfg: AutoloadSettings) -> PhotoNamingSettings:
    return PhotoNamingSettings(
        yandex_disk_root=cfg.yandex_disk_root,
        image_count=cfg.image_count,
        image_ext=cfg.image_ext,
        photo_layout=getattr(cfg, "photo_layout", "flat"),
        photos_public_base_url=getattr(cfg, "photos_public_base_url", ""),
    )


MAX_AVITO_DESCRIPTION_LEN = 7500
# В шаблоне Авито колонка «Количество» = за сколько шт указана «Цена», не остаток на складе.
AUTOLOAD_PRICE_QUANTITY = "1"


class _SafeDict(dict):
    def __missing__(self, key):
        return ""


def _availability_headline(ushk_in_stock: bool) -> str:
    if ushk_in_stock:
        return "Шины в наличии!"
    return "Шины под заказ 1-2 дня"


def _posting_ushk_in_stock(post_row) -> bool:
    if post_row is None:
        return False
    val = post_row.get("ушк_в_наличии")
    if val is True:
        return True
    return str(val or "").strip().lower() in ("true", "1", "да", "yes")


def _posting_sam_mb_cash_price(post_row) -> bool:
    if post_row is None:
        return False
    val = post_row.get("цена_за_наличный_расчет")
    if val is True:
        return True
    return str(val or "").strip().lower() in ("true", "1", "да", "yes")


def _payment_terms(sam_mb_cash_price: bool) -> str:
    if sam_mb_cash_price:
        return "Цена за наличный расчет"
    return "Любая форма оплаты, НДС"


def _autoload_price(value) -> int:
    """Цена в Excel автозагрузки — до десятков рублей."""
    return round_price_to_tens(float(value))


def _apply_prices_to_sheet(
    ws,
    *,
    price_col: int | None,
    title_col: int | None,
    posting_df: pd.DataFrame,
) -> int:
    """
    Все строки данных: цена из posting (если есть номенклатура), иначе округление текущей.
    Нужно и для строк без фото — в шаблоне оставались старые значения вроде 4108.
    """
    if not price_col:
        return 0

    post_by_nom: dict[str, int] = {}
    for _, row in posting_df.iterrows():
        nom = str(row.get("номенклатура", "") or "").strip()
        rec = row.get("recommended_price")
        if nom and not pd.isna(rec):
            post_by_nom[nom] = _autoload_price(rec)

    changed = 0
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        title = ""
        if title_col:
            title = str(ws.cell(row_idx, title_col).value or "").strip()
        if title and title in post_by_nom:
            new_price = post_by_nom[title]
        else:
            raw = ws.cell(row_idx, price_col).value
            if raw in (None, "") or str(raw).strip() == "":
                continue
            try:
                new_price = _autoload_price(raw)
            except (TypeError, ValueError):
                continue
        current = ws.cell(row_idx, price_col).value
        try:
            same = current is not None and int(float(current)) == new_price
        except (TypeError, ValueError):
            same = False
        if not same:
            ws.cell(row_idx, price_col, new_price)
            changed += 1
    return changed


def _apply_quantities_to_sheet(
    ws,
    *,
    qty_col: int | None,
    title_col: int | None,
    posting_df: pd.DataFrame | None = None,
    max_quantity: int | None = None,
) -> int:
    """Колонка «Количество» = всегда 1 (цена в файле за 1 шт). Остаток — только в описании."""
    del posting_df, max_quantity
    if not qty_col:
        return 0

    changed = 0
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        if not title_col:
            break
        title = str(ws.cell(row_idx, title_col).value or "").strip()
        if not title:
            continue
        current = str(ws.cell(row_idx, qty_col).value or "").strip()
        if current != AUTOLOAD_PRICE_QUANTITY:
            ws.cell(row_idx, qty_col, AUTOLOAD_PRICE_QUANTITY)
            changed += 1
    return changed


def _apply_descriptions_to_sheet(
    ws,
    *,
    desc_col: int | None,
    title_col: int | None,
    price_col: int | None,
    id_col: int | None,
    qty_col: int | None,
    posting_df: pd.DataFrame,
    cfg: AutoloadSettings,
    stores: StoresConfig,
    model_descriptions: dict[str, str],
) -> int:
    """Все строки с названием: описание по шаблону (в т.ч. без фото)."""
    if not desc_col or not title_col:
        return 0

    post_by_nom = _posting_row_lookup(
        posting_df,
        max_quantity=cfg.max_listing_quantity,
    )
    changed = 0

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        nom = str(ws.cell(row_idx, title_col).value or "").strip()
        if not nom:
            continue

        listing_id = (
            str(ws.cell(row_idx, id_col).value or "").strip() if id_col else ""
        )
        article = _article_from_listing_id(listing_id)
        post = post_by_nom.get(nom, {})
        if post.get("article"):
            article = post["article"]

        if post.get("price") is not None:
            price_int = post["price"]
        elif price_col:
            raw_price = ws.cell(row_idx, price_col).value
            try:
                price_int = (
                    _autoload_price(raw_price)
                    if raw_price not in (None, "")
                    else 0
                )
            except (TypeError, ValueError):
                price_int = 0
        else:
            price_int = 0

        if post.get("quantity"):
            stock_qty = post["quantity"]
        else:
            stock_qty = "1"

        fields = parse_title_fields(nom)
        model_desc = lookup_model_description(
            model_descriptions,
            nomenclature=nom,
            brand=fields.get("brand", ""),
            model=fields.get("model", ""),
        )
        row_defaults = _store_defaults_for_listing_id(
            listing_id, stores, cfg.defaults
        )
        new_desc = _format_description(
            cfg.description_html,
            nomenclature=nom,
            article=article,
            price=price_int,
            quantity=stock_qty,
            model_description=model_desc,
            store_pitch=cfg.store_pitch_html,
            store_defaults=row_defaults,
            ushk_in_stock=bool(post.get("ushk_in_stock")),
            sam_mb_cash_price=bool(post.get("sam_mb_cash_price")),
        )
        current = str(ws.cell(row_idx, desc_col).value or "")
        if current != new_desc:
            ws.cell(row_idx, desc_col, new_desc)
            changed += 1

    return changed


def _prototype_clear_cols(
    *,
    photos_col: int | None,
    avito_col: int | None,
) -> frozenset[int]:
    """Колонки, которые нельзя копировать из первой строки шаблона."""
    cols: set[int] = set()
    if photos_col:
        cols.add(photos_col)
    if avito_col:
        cols.add(avito_col)
    return frozenset(cols)


def _append_row_from_prototype(
    ws,
    row_idx: int,
    prototype_row: list,
    *,
    clear_cols: frozenset[int],
) -> None:
    for c in range(1, len(prototype_row) + 1):
        val = "" if c in clear_cols else copy.copy(prototype_row[c - 1])
        ws.cell(row_idx, c, val)


def _avito_id_for_row(
    listing_id: str,
    article: str,
    avito_ids: dict[str, str],
) -> str:
    if listing_id and listing_id in avito_ids:
        return avito_ids[listing_id]
    if article and article in avito_ids:
        return avito_ids[article]
    return ""


def _sync_avito_ids_to_sheet(
    ws,
    *,
    avito_col: int | None,
    id_col: int | None,
    title_col: int | None,
    posting_df: pd.DataFrame,
    avito_ids: dict[str, str],
) -> tuple[int, int]:
    """
    Номер объявления на Авито — только свой для каждой строки.

    Иначе из шаблона копируется один номер на десятки md_артикулов.
    """
    if not avito_col:
        return 0, 0

    post_by_nom = _posting_row_lookup(posting_df)
    set_count = 0
    cleared = 0

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        listing_id = ""
        if id_col:
            listing_id = normalize_article_id(ws.cell(row_idx, id_col).value)
        article = _article_from_listing_id(listing_id) if listing_id else ""
        if not article and title_col:
            nom = str(ws.cell(row_idx, title_col).value or "").strip()
            article = post_by_nom.get(nom, {}).get("article", "")
            if article and not listing_id:
                listing_id = article

        expected = _avito_id_for_row(listing_id, article, avito_ids)
        if expected:
            expected = expected.split(".")[0]

        current = str(ws.cell(row_idx, avito_col).value or "").strip()
        if current.endswith(".0"):
            current = current.split(".")[0]
        if current.lower() == "nan":
            current = ""

        if expected:
            if current != expected:
                ws.cell(row_idx, avito_col, expected)
                set_count += 1
        elif current:
            ws.cell(row_idx, avito_col, "")
            cleared += 1

    return set_count, cleared


def _resolve_yandex_downloader(
    cfg: AutoloadSettings,
    project_root: Path,
    secrets_file: Path | None,
) -> YandexDiskDownloadUrls | None:
    if cfg.image_mode != "yandex_https":
        return None
    sec = secrets_file or (project_root / "secrets.local.yaml")
    if not sec.is_absolute():
        sec = project_root / sec
    token = load_yandex_oauth_token(sec)
    return YandexDiskDownloadUrls(token)


def _should_replace_photo_urls(current: str, *, source: str) -> bool:
    """
    Фото артикула всегда пишем в файл (замена модели или ссылок Avito).
    Модель — не трогаем уже принятые Avito-ссылки.
    """
    if source == "article":
        return True
    return not is_avito_hosted_photo_urls(current)


def _photo_urls_for_article(
    local_photos: Path | None,
    article: str,
    stores: StoresConfig,
    cfg: AutoloadSettings,
    *,
    brand: str = "",
    model: str = "",
    yandex_downloader: YandexDiskDownloadUrls | None = None,
) -> tuple[str, str]:
    if not local_photos or not article:
        return "", ""
    resolved = resolve_listing_photo_sets(
        local_photos,
        article,
        stores.prefixes,
        layout=cfg.photo_layout,
        prefix_in_filename=cfg.photo_store_prefix_in_filename,
        brand=brand,
        model=model,
        model_fallback=cfg.model_photo_fallback,
        article_first=cfg.photo_article_first,
        legacy_unprefixed_prefix=stores.legacy_unprefixed_store,
        max_count=int(cfg.image_count or 0),
        jpeg_quality=cfg.jpeg_quality,
        contributors_prefix=cfg.contributors_prefix,
    )
    if not resolved.store_sets:
        return "", ""
    sp = resolved.store_sets[0]
    urls = build_store_photo_urls(
        sp,
        _photo_cfg(cfg),
        article=article,
        layout=cfg.photo_layout,
        image_mode=cfg.image_mode,
        photos_root=local_photos,
        downloader=yandex_downloader,
    )
    if cfg.image_mode == "yandex_disk":
        return normalize_yandex_photo_urls(urls), resolved.source
    return urls, resolved.source


def _article_for_sheet_row(
    ws,
    row_idx: int,
    *,
    id_col: int | None,
    title_col: int | None,
    post_by_nom: dict[str, dict],
) -> str:
    if id_col:
        listing_id = normalize_article_id(ws.cell(row_idx, id_col).value)
        if listing_id and "_" in listing_id:
            return _article_from_listing_id(listing_id)
    if title_col:
        nom = str(ws.cell(row_idx, title_col).value or "").strip()
        post = post_by_nom.get(nom, {})
        if post.get("article"):
            return post["article"]
    return ""


def _sync_photo_urls_to_sheet(
    ws,
    *,
    photos_col: int | None,
    title_col: int | None,
    id_col: int | None,
    posting_df: pd.DataFrame,
    local_photos: Path | None,
    cfg: AutoloadSettings,
    stores: StoresConfig,
    yandex_downloader: YandexDiskDownloadUrls | None = None,
) -> tuple[int, int]:
    """
    Фото только если файлы есть на Диске; иначе ячейку очищаем.

    Без этого при копировании prototype в новые строки тянутся чужие ссылки.
    Ссылки Avito не перезаписываем, кроме замены на фото артикула.
    """
    if not photos_col:
        return 0, 0

    post_by_nom = _posting_row_lookup(posting_df)
    set_count = 0
    cleared = 0

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        current = str(ws.cell(row_idx, photos_col).value or "").strip()

        article = _article_for_sheet_row(
            ws,
            row_idx,
            id_col=id_col,
            title_col=title_col,
            post_by_nom=post_by_nom,
        )
        brand = model = ""
        if title_col:
            nom = str(ws.cell(row_idx, title_col).value or "").strip()
            if nom:
                parsed = parse_title_fields(nom)
                brand = parsed.get("brand", "")
                model = parsed.get("model", "")
        photos, source = _photo_urls_for_article(
            local_photos,
            article,
            stores,
            cfg,
            brand=brand,
            model=model,
            yandex_downloader=yandex_downloader,
        )

        if photos:
            if current != photos and _should_replace_photo_urls(current, source=source):
                ws.cell(row_idx, photos_col, photos)
                set_count += 1
        elif current and _should_replace_photo_urls(current, source=""):
            ws.cell(row_idx, photos_col, "")
            cleared += 1

    return set_count, cleared


def _format_price(value) -> str:
    try:
        return format(_autoload_price(value), "_").replace("_", " ")
    except Exception:
        return ""


def _article_from_listing_id(listing_id: str) -> str:
    sid = normalize_article_id(listing_id)
    if "_" in sid:
        return sid.split("_", 1)[1]
    return sid


def _store_defaults_for_listing_id(
    listing_id: str,
    stores: StoresConfig,
    cfg_defaults: dict[str, str],
) -> dict[str, str]:
    sid = normalize_article_id(listing_id)
    prefix = ""
    if "_" in sid:
        prefix = sid.split("_", 1)[0]
    elif stores.legacy_unprefixed_store:
        prefix = stores.legacy_unprefixed_store
    store = stores.get(prefix)
    if store:
        return merge_defaults(cfg_defaults, store)
    return dict(cfg_defaults)


def _posting_row_lookup(
    posting_df: pd.DataFrame,
    *,
    max_quantity: int = 12,
) -> dict[str, dict]:
    out: dict[str, dict] = {}
    for _, row in posting_df.iterrows():
        nom = str(row.get("номенклатура", "") or "").strip()
        if not nom:
            continue
        rec = row.get("recommended_price")
        if pd.isna(rec):
            continue
        out[nom] = {
            "article": normalize_article_id(row.get("артикул", "")),
            "price": _autoload_price(rec),
            "quantity": _quantity_label(
                str(row.get("количество", "")),
                max_quantity=max_quantity,
            ),
            "ushk_in_stock": _posting_ushk_in_stock(row),
            "sam_mb_cash_price": _posting_sam_mb_cash_price(row),
        }
    return out


def _format_description(
    template: str,
    *,
    nomenclature: str,
    article: str,
    price: int,
    quantity: str,
    model_description: str,
    store_pitch: str = "",
    store_defaults: dict[str, str],
    ushk_in_stock: bool = False,
    sam_mb_cash_price: bool = False,
) -> str:
    payload = _SafeDict(
        nomenclature=nomenclature,
        article=article,
        price=str(price),
        price_human=_format_price(price),
        quantity=quantity,
        availability_headline=_availability_headline(ushk_in_stock),
        payment_terms=_payment_terms(sam_mb_cash_price),
        model_description=model_description,
        store_pitch=store_pitch or "",
        contact_person=store_defaults.get("contact_person", ""),
        phone=store_defaults.get("phone", ""),
        address=store_defaults.get("address", ""),
        company=store_defaults.get("company", ""),
        email=store_defaults.get("email", ""),
        contact_method=store_defaults.get("contact_method", ""),
    )
    desc = template.format_map(payload)
    return desc[:MAX_AVITO_DESCRIPTION_LEN]


def _quantity_label(qty: str, *, max_quantity: int = 12) -> str:
    """Остаток на складе для текста описания (не для колонки «Количество» в Excel)."""
    q = str(qty).strip()
    if not q or q.lower() == "nan":
        return "1"
    try:
        n = int(float(q))
        if n <= 0:
            return "1"
        cap = max(1, int(max_quantity))
        return str(min(n, cap))
    except ValueError:
        return "1"


def _to_float_or_none(value) -> float | None:
    if value is None:
        return None
    s = str(value).strip().replace(" ", "").replace(",", ".")
    if not s or s.lower() in ("nan", "none"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _is_priority_for_photo_queue(post_row) -> bool:
    """
    Приоритет для no_photos:
    - товара нет на Avito, либо
    - наша рекомендованная цена ниже avito_min.
    """
    on_avito = str(post_row.get("есть_на_avito", "")).strip().lower() in (
        "true",
        "1",
        "да",
    )
    if not on_avito:
        return True
    rec = _to_float_or_none(post_row.get("recommended_price"))
    avito_min = _to_float_or_none(post_row.get("avito_min"))
    if rec is None or avito_min is None:
        return False
    return rec < avito_min


def load_posting(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="к выкладке")
    return df


def find_latest_avito_export(search_dir: Path) -> Path | None:
    """Последний скачанный с Авито xlsx (432801655_*.xlsx)."""
    if not search_dir.is_dir():
        return None
    files = sorted(
        search_dir.glob("432801655_*.xlsx"),
        key=lambda p: p.stat().st_mtime,
    )
    return files[-1] if files else None


def resolve_autoload_base(
    cfg: AutoloadSettings,
    *,
    root: Path,
    override: Path | None = None,
    use_working: bool = False,
) -> tuple[Path, str]:
    """
    База для автозагрузки.

    По умолчанию: последняя выгрузка 432801655_*.xlsx в папке template_file.
    --use-working: накопленный autoload_working.xlsx.
    """
    if override:
        return override, "аргумент --template"

    if use_working:
        working = root / cfg.working_file
        if working.exists():
            return working, f"накопленный {cfg.working_file}"

    if cfg.prefer_latest_avito_export:
        template_dir = (root / cfg.template_file).parent
        latest = find_latest_avito_export(template_dir)
        if latest:
            return latest, f"последняя выгрузка Авито ({latest.name})"

    template = root / cfg.template_file
    if template.exists():
        return template, f"шаблон {cfg.template_file}"

    working = root / cfg.working_file
    if working.exists():
        return working, f"накопленный {cfg.working_file} (fallback)"

    return template, f"шаблон {cfg.template_file}"


def extract_avito_ids_from_xlsx(path: Path, stores: StoresConfig | None = None) -> dict[str, str]:
    """
    Номера объявлений из выгрузки Авито.

    Поддерживает оба формата Id:
    - наш: md_12044 в колонке «Уникальный идентификатор»;
    - выгрузка Авито: числовой Id + сопоставление по «Название объявления».
    """
    by_listing, _by_title = _extract_avito_export_maps(path)
    return by_listing


def _extract_avito_export_maps(
    path: Path,
) -> tuple[dict[str, str], dict[str, str]]:
    if not path.exists():
        return {}, {}
    # read_only ломает max_row/заголовки на выгрузках Авито — грузим обычно
    wb = load_workbook(path, data_only=True)
    try:
        sheet_name = _find_data_sheet(wb)
        ws = wb[sheet_name]
        headers = _header_map(ws)
        c_id = _col(headers, "Уникальный идентификатор объявления")
        c_avito = _col(headers, "Номер объявления на Авито")
        c_title = _col(headers, "Название объявления")
        if not c_avito:
            return {}, {}
        by_listing: dict[str, str] = {}
        by_title: dict[str, str] = {}
        for row in range(DATA_START_ROW, ws.max_row + 1):
            avito_num = str(ws.cell(row, c_avito).value or "").strip()
            if not avito_num or avito_num.lower() == "nan":
                continue
            avito_num = avito_num.split(".")[0]
            title = (
                str(ws.cell(row, c_title).value or "").strip() if c_title else ""
            )
            if title:
                by_title[title] = avito_num
            if not c_id:
                continue
            listing_id = normalize_article_id(ws.cell(row, c_id).value)
            if not listing_id:
                continue
            if "_" in listing_id:
                by_listing[listing_id] = avito_num
                article = _article_from_listing_id(listing_id)
                if article:
                    by_listing[article] = avito_num
            elif title:
                by_listing[f"title:{title}"] = avito_num
        return by_listing, by_title
    finally:
        wb.close()


def avito_ids_for_posting(
    posting_df: pd.DataFrame,
    stores: StoresConfig,
    *,
    ids_from_xlsx: dict[str, str] | None = None,
    titles_from_xlsx: dict[str, str] | None = None,
    ids_from_csv: dict[str, str] | None = None,
) -> dict[str, str]:
    """Собрать avito_id по артикулу / md_артикул / названию из posting."""
    by_listing = dict(ids_from_xlsx or {})
    by_title = dict(titles_from_xlsx or {})
    for key, val in list(by_listing.items()):
        if key.startswith("title:"):
            by_title[key[6:]] = val
    out = merge_avito_ids(by_listing, ids_from_csv or {}, stores=stores)

    for _, row in posting_df.iterrows():
        nom = str(row.get("номенклатура", "") or "").strip()
        art = normalize_article_id(row.get("артикул", ""))
        avito_num = ""
        if nom and nom in by_title:
            avito_num = by_title[nom]
        elif nom and f"title:{nom}" in by_listing:
            avito_num = by_listing[f"title:{nom}"]
        if not avito_num or not art:
            continue
        out[art] = avito_num
        for store in stores.stores:
            out[store.listing_id(art)] = avito_num
    return out


def merge_avito_ids(
    *maps: dict[str, str],
    stores: StoresConfig | None = None,
) -> dict[str, str]:
    """Поздние словари перекрывают ранние; дублируем по префиксам магазинов."""
    out: dict[str, str] = {}
    store_map = stores.by_prefix() if stores else {}
    for m in maps:
        for key, val in m.items():
            if not key or not val:
                continue
            out[key] = val
            if "_" in key:
                prefix, art = key.split("_", 1)
                if prefix in store_map and art:
                    out[art] = val
    return out


def save_avito_ids_csv(path: Path, mapping: dict[str, str]) -> int:
    """Записать avito_ids.csv (артикул;avito_id), без md_ ключей."""
    rows: list[tuple[str, str]] = []
    seen: set[str] = set()
    for key, val in sorted(mapping.items()):
        if "_" in key:
            continue
        if key in seen:
            continue
        seen.add(key)
        rows.append((key, val))
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = ["артикул;avito_id", *(f"{a};{i}" for a, i in rows)]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8-sig")
    return len(rows)


def load_avito_ids(path: Path, stores: StoresConfig | None = None) -> dict[str, str]:
    if not path.exists():
        return {}
    df = pd.read_csv(path, encoding="utf-8-sig", sep=None, engine="python")
    cols = {str(c).strip().lower(): c for c in df.columns}
    art = cols.get("артикул") or cols.get("article")
    aid = cols.get("avito_id") or cols.get("номер объявления на авито") or cols.get("avito id")
    prefix_col = cols.get("префикс") or cols.get("prefix") or cols.get("магазин")
    if not art or not aid:
        return {}
    out: dict[str, str] = {}
    store_map = stores.by_prefix() if stores else {}
    for _, r in df.iterrows():
        a = normalize_article_id(r[art])
        i = str(r[aid]).strip()
        if not a or not i or i.lower() == "nan":
            continue
        avito_id = i.split(".")[0]
        out[a] = avito_id
        if prefix_col:
            p = str(r[prefix_col]).strip()
            if p and p in store_map:
                out[store_map[p].listing_id(a)] = avito_id
        elif stores:
            for store in stores.stores:
                if a.startswith(f"{store.prefix}_"):
                    out[a] = avito_id
    return out


def _listing_id_for_article(
    article: str,
    stores: StoresConfig,
    *,
    current_id: str = "",
) -> str:
    """md_12044 — наш Id для автозагрузки (не числовой Id Авито)."""
    prefix = ""
    if current_id and "_" in current_id:
        p = current_id.split("_", 1)[0]
        if p in stores.by_prefix():
            prefix = p
    if not prefix and stores.stores:
        prefix = stores.stores[0].prefix
    store = stores.get(prefix) if prefix else None
    if store:
        return store.listing_id(article)
    return f"{prefix}_{article}" if prefix else article


def _set_unique_id(ws, row_idx: int, col: int | None, listing_id: str) -> None:
    if not col or not listing_id:
        return
    current = normalize_article_id(ws.cell(row_idx, col).value)
    if current != listing_id:
        ws.cell(row_idx, col, listing_id)


def _apply_listing_ids_to_sheet(
    ws,
    *,
    id_col: int | None,
    title_col: int | None,
    posting_df: pd.DataFrame,
    stores: StoresConfig,
) -> int:
    """
    Заменить числовые Id из выгрузки Авито на md_артикул.

    Иначе все строки имеют один Id, индекс by_id ломается, обновляется одна строка.
    """
    if not id_col or not title_col:
        return 0

    post_by_nom: dict[str, str] = {}
    for _, row in posting_df.iterrows():
        nom = str(row.get("номенклатура", "") or "").strip()
        art = normalize_article_id(row.get("артикул", ""))
        if nom and art:
            post_by_nom[nom] = art

    changed = 0
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        nom = str(ws.cell(row_idx, title_col).value or "").strip()
        if not nom:
            continue
        art = post_by_nom.get(nom)
        if not art:
            continue
        current = normalize_article_id(ws.cell(row_idx, id_col).value)
        listing_id = _listing_id_for_article(art, stores, current_id=current)
        if current != listing_id:
            ws.cell(row_idx, id_col, listing_id)
            changed += 1
    return changed


def _ensure_posting_rows_in_sheet(
    ws,
    *,
    posting_df: pd.DataFrame,
    c_id: int | None,
    c_title: int | None,
    c_photos: int | None,
    c_avito: int | None,
    stores: StoresConfig,
    prototype_row: list | None,
    by_title: dict[str, int],
    by_id: dict[str, int],
    next_row: int,
    stats: dict[str, int],
) -> int:
    """
    Добавить в лист все позиции из posting, которых ещё нет.

    Выгрузка Авито содержит только активные объявления (~30); остаток goods
    дописываем сюда. Без фото строка всё равно нужна — цена/описание обновятся.
    """
    if prototype_row is None:
        return 0

    added = 0
    for _, post in posting_df.iterrows():
        if post.get("дубликат_остаток") is True or str(
            post.get("дубликат_остаток")
        ).lower() == "true":
            continue
        nom = str(post.get("номенклатура", "") or "").strip()
        if not nom or pd.isna(post.get("recommended_price")):
            continue
        if nom in by_title:
            continue
        article = normalize_article_id(post.get("артикул", ""))
        if not article:
            continue

        listing_id = _listing_id_for_article(article, stores)
        row_idx = next_row
        _append_row_from_prototype(
            ws,
            row_idx,
            prototype_row,
            clear_cols=_prototype_clear_cols(
                photos_col=c_photos, avito_col=c_avito
            ),
        )
        _set_unique_id(ws, row_idx, c_id, listing_id)
        if c_title:
            ws.cell(row_idx, c_title, nom)
        by_title[nom] = row_idx
        by_id[listing_id] = row_idx
        by_id[article] = row_idx
        next_row += 1
        added += 1
        stats["appended"] += 1

    return added


def _set_avito_id(
    ws,
    row_idx: int,
    col: int | None,
    listing_id: str,
    avito_ids: dict[str, str],
) -> None:
    if not col or not listing_id:
        return
    article = _article_from_listing_id(listing_id)
    expected = _avito_id_for_row(listing_id, article, avito_ids)
    if not expected:
        return
    current = str(ws.cell(row_idx, col).value or "").strip().split(".")[0]
    if current != expected:
        ws.cell(row_idx, col, expected)


def _remove_rows_without_photos(
    ws,
    *,
    photos_col: int | None,
    title_col: int | None,
) -> int:
    """Убрать из файла автозагрузки строки без фото на Диске."""
    if not photos_col:
        return 0
    to_delete: list[int] = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        photos = str(ws.cell(row, photos_col).value or "").strip()
        if photos:
            continue
        title = ""
        if title_col:
            title = str(ws.cell(row, title_col).value or "").strip()
        if not title:
            continue
        to_delete.append(row)
    for row in reversed(to_delete):
        ws.delete_rows(row, 1)
    return len(to_delete)


def _remove_rows_not_in_goods(
    ws,
    *,
    c_id: int | None,
    c_title: int | None,
    c_avito: int | None,
    keep_articles: set[str],
    keep_titles: set[str],
    keep_listing_ids: set[str],
) -> list[dict]:
    """Удалить строки с объявлением, которых нет в goods (→ архив на Авито)."""
    to_delete: list[tuple[int, dict]] = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        title = ""
        if c_title:
            title = str(ws.cell(row, c_title).value or "").strip()
        article = normalize_article_id(ws.cell(row, c_id).value if c_id else None)
        if not title and not article:
            continue
        row_id = normalize_article_id(ws.cell(row, c_id).value if c_id else None)
        if row_in_goods(
            row_id=row_id,
            article=article,
            title=title,
            keep_articles=keep_articles,
            keep_titles=keep_titles,
            keep_listing_ids=keep_listing_ids,
        ):
            continue
        avito_num = ""
        if c_avito:
            avito_num = str(ws.cell(row, c_avito).value or "").strip()
        to_delete.append(
            (
                row,
                {
                    "строка": row,
                    "артикул": article,
                    "название": title,
                    "avito_id": avito_num,
                },
            )
        )

    removed = [info for _, info in to_delete]
    for row, _ in reversed(to_delete):
        ws.delete_rows(row, 1)
    return removed


_STORE_FORCE_KEYS = frozenset(
    {"contact_person", "phone", "address", "contact_method", "company", "email"}
)

# MultiItem / «Мультиобъявление» в ЛК Авито
_MERGE_ADS_HEADERS = (
    "Соединять это объявление с другими объявлениями",
    "Мультиобъявление",
    "MultiItem",
)

_MULTI_NAME_HEADERS = (
    "Название мультиобъявления",
    "MultiName",
)


def _col_any(headers: dict[str, int], names: tuple[str, ...]) -> int | None:
    for name in names:
        col = _col(headers, name)
        if col:
            return col
    return None


def _sync_merge_ads_to_sheet(
    ws,
    headers: dict[str, int],
    *,
    value: str = "Да",
) -> int:
    """Включить мультиобъявление (MultiItem) во всех строках данных."""
    col = _col_any(headers, _MERGE_ADS_HEADERS)
    if not col:
        return 0
    changed = 0
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        current = str(ws.cell(row_idx, col).value or "").strip()
        if current.lower() in ("", "nan", "нет", "no", "false", "0"):
            ws.cell(row_idx, col, value)
            changed += 1
        elif current != value:
            ws.cell(row_idx, col, value)
            changed += 1
    return changed


def _sync_multi_names_to_sheet(
    ws,
    headers: dict[str, int],
    *,
    title_col: int | None,
) -> tuple[int, int]:
    """MultiName по размеру из названия (все 195/65 R15 → 19565R15)."""
    col = _col_any(headers, _MULTI_NAME_HEADERS)
    if not col or not title_col:
        return 0, 0
    changed = 0
    groups: set[str] = set()
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        title = str(ws.cell(row_idx, title_col).value or "").strip()
        if not title:
            continue
        multi_name = build_multi_name_from_title(title)
        current = str(ws.cell(row_idx, col).value or "").strip()
        if current.endswith(".0"):
            current = current.split(".")[0]
        if current.lower() == "nan":
            current = ""
        if multi_name:
            groups.add(multi_name)
            if current != multi_name:
                ws.cell(row_idx, col, multi_name)
                changed += 1
        elif current:
            ws.cell(row_idx, col, "")
            changed += 1
    return changed, len(groups)


def _apply_defaults(
    ws,
    row_idx: int,
    headers: dict[str, int],
    defaults: dict[str, str],
    *,
    skip_keys: frozenset[str] = frozenset({"run_flat", "condition"}),
    force_keys: frozenset[str] = frozenset(),
) -> None:
    for key, val in defaults.items():
        col = _col(headers, _DEFAULT_HEADER_ALIASES.get(key, key))
        if col and val and key not in skip_keys:
            current = ws.cell(row_idx, col).value
            if key in force_keys or current in (None, ""):
                ws.cell(row_idx, col, val)


def fill_autoload_template(
    *,
    template_path: Path,
    posting_df: pd.DataFrame,
    cfg: AutoloadSettings,
    stores: StoresConfig,
    model_descriptions: dict[str, str],
    avito_ids: dict[str, str],
    output_path: Path,
    project_root: Path | None = None,
    secrets_file: Path | None = None,
) -> dict[str, int]:
    wb = load_workbook(template_path)
    sheet_name = cfg.sheet_name or _find_data_sheet(wb)
    ws = wb[sheet_name]
    headers = _header_map(ws)

    c_id = _col(headers, "Уникальный идентификатор объявления")
    c_avito = _col(headers, "Номер объявления на Авито")
    c_title = _col(headers, "Название объявления")
    c_price = _col(headers, "Цена")
    c_photos = _col(headers, "Ссылки на фото")
    c_desc = _col(headers, "Описание объявления")
    c_qty = _col(headers, "Количество")
    c_brand = _col(headers, "Производитель")
    c_model = _col(headers, "Модель")
    c_w = _col(headers, "Ширина профиля")
    c_p = _col(headers, "Высота профиля")
    c_d = _col(headers, "Диаметр")
    c_li = _col(headers, "Индекс нагрузки")
    c_si = _col(headers, "Индекс скорости")
    c_season = _col(headers, "Сезонность")
    c_run = _col(headers, "Run Flat")
    c_state = _col(headers, "Состояние")

    required = [c_id, c_title, c_price]
    if any(x is None for x in required):
        raise ValueError(f"В шаблоне не найдены обязательные колонки. Есть: {list(headers)[:8]}...")

    root = project_root or Path.cwd()
    local_photos = resolve_photos_folder(cfg, root)
    yandex_downloader = _resolve_yandex_downloader(cfg, root, secrets_file)
    photos_missing: list[dict] = []
    missing_models: dict[str, dict] = {}

    listing_ids_fixed = _apply_listing_ids_to_sheet(
        ws,
        id_col=c_id,
        title_col=c_title,
        posting_df=posting_df,
        stores=stores,
    )
    if listing_ids_fixed:
        LOG.info(
            "Уникальный Id: исправлено %s строк (md_артикул вместо Id Авито)",
            listing_ids_fixed,
        )

    keep_articles, keep_titles, keep_listing_ids = posting_keep_sets(
        posting_df, stores
    )
    removed: list[dict] = []
    if cfg.close_not_in_goods:
        removed = _remove_rows_not_in_goods(
            ws,
            c_id=c_id,
            c_title=c_title,
            c_avito=c_avito,
            keep_articles=keep_articles,
            keep_titles=keep_titles,
            keep_listing_ids=keep_listing_ids,
        )

    # Пересобрать индекс после удалений
    by_title: dict[str, int] = {}
    by_id: dict[str, int] = {}
    prototype_row: list | None = None
    for row in range(DATA_START_ROW, ws.max_row + 1):
        title_val = ws.cell(row, c_title).value if c_title else None
        id_val = ws.cell(row, c_id).value if c_id else None
        if prototype_row is None and title_val:
            prototype_row = [ws.cell(row, c).value for c in range(1, ws.max_column + 1)]
        if title_val:
            by_title[str(title_val).strip()] = row
        if id_val:
            sid = normalize_article_id(id_val)
            if sid and "_" in sid:
                by_id[sid] = row
                art = _article_from_listing_id(sid)
                if art:
                    by_id[art] = row

    stats: dict[str, int] = {
        "updated": 0,
        "appended": 0,
        "skipped": 0,
        "skipped_no_photos": 0,
        "model_photo_fallback": 0,
        "removed": len(removed),
    }
    next_row = ws.max_row + 1
    if next_row < DATA_START_ROW:
        next_row = DATA_START_ROW

    if cfg.include_all_goods_in_autoload:
        added_from_posting = _ensure_posting_rows_in_sheet(
            ws,
            posting_df=posting_df,
            c_id=c_id,
            c_title=c_title,
            c_photos=c_photos,
            c_avito=c_avito,
            stores=stores,
            prototype_row=prototype_row,
            by_title=by_title,
            by_id=by_id,
            next_row=next_row,
            stats=stats,
        )
        if added_from_posting:
            LOG.info(
                "Из posting добавлено %s строк (режим полного goods)",
                added_from_posting,
            )

    for _, post in posting_df.iterrows():
        if post.get("дубликат_остаток") is True or str(post.get("дубликат_остаток")).lower() == "true":
            stats["skipped"] += 1
            continue

        nom = str(post.get("номенклатура", "")).strip()
        if not nom:
            stats["skipped"] += 1
            continue

        article = normalize_article_id(post.get("артикул", ""))
        price = post.get("recommended_price")
        if pd.isna(price):
            stats["skipped"] += 1
            continue
        price_int = _autoload_price(price)

        fields = parse_title_fields(nom)
        resolved_photos = resolve_listing_photo_sets(
            local_photos,
            article,
            stores.prefixes,
            layout=cfg.photo_layout,
            prefix_in_filename=cfg.photo_store_prefix_in_filename,
            brand=fields.get("brand", ""),
            model=fields.get("model", ""),
            model_fallback=cfg.model_photo_fallback,
            article_first=cfg.photo_article_first,
            legacy_unprefixed_prefix=stores.legacy_unprefixed_store,
            max_count=int(cfg.image_count or 0),
            jpeg_quality=cfg.jpeg_quality,
            contributors_prefix=cfg.contributors_prefix,
        )
        store_photo_sets = list(resolved_photos.store_sets)
        if resolved_photos.source == "model":
            stats["model_photo_fallback"] += 1
        if cfg.skip_without_photos:
            reason = ""
            if cfg.verify_photos_on_disk and not local_photos:
                reason = "папка фото не найдена"
            elif not store_photo_sets:
                reason = "нет фото на диске (артикул и модель)"
            if reason:
                if _is_priority_for_photo_queue(post):
                    photos_missing.append(
                        {
                            "артикул": article,
                            "номенклатура": nom,
                            "магазины": ", ".join(stores.prefixes),
                            "проблема": reason,
                        }
                    )
                stats["skipped"] += 1
                stats["skipped_no_photos"] += 1
                continue

        model_key = " ".join(
            x for x in (fields.get("brand", ""), fields.get("model", "")) if x
        ).strip()
        model_desc = lookup_model_description(
            model_descriptions,
            nomenclature=nom,
            brand=fields.get("brand", ""),
            model=fields.get("model", ""),
        )
        if model_key and not model_desc:
            missing_models[model_key] = {
                "модель": model_key,
                "бренд": fields.get("brand", ""),
                "model": fields.get("model", ""),
                "пример_номенклатуры": nom,
            }
        stock_qty = _quantity_label(
            str(post.get("количество", "")),
            max_quantity=cfg.max_listing_quantity,
        )
        photo_cfg = _photo_cfg(cfg)

        for sp in store_photo_sets:
            store = stores.get(sp.prefix)
            if not store:
                continue

            listing_id = store.listing_id(article)
            photos = build_store_photo_urls(
                sp,
                photo_cfg,
                article=article,
                layout=cfg.photo_layout,
                image_mode=cfg.image_mode,
                photos_root=local_photos,
                downloader=yandex_downloader,
            )
            if cfg.image_mode == "yandex_disk":
                photos = normalize_yandex_photo_urls(photos)
            row_defaults = merge_defaults(cfg.defaults, store)

            row_idx = by_id.get(listing_id)
            if row_idx is None and article:
                row_idx = by_id.get(article)

            if row_idx is None:
                if prototype_row is None:
                    stats["skipped"] += 1
                    continue
                row_idx = next_row
                _append_row_from_prototype(
                    ws,
                    row_idx,
                    prototype_row,
                    clear_cols=_prototype_clear_cols(
                        photos_col=c_photos, avito_col=c_avito
                    ),
                )
                next_row += 1
                stats["appended"] += 1
                by_id[listing_id] = row_idx
            else:
                stats["updated"] += 1

            _set_unique_id(ws, row_idx, c_id, listing_id)
            _set_avito_id(ws, row_idx, c_avito, listing_id, avito_ids)
            ws.cell(row_idx, c_title, nom)
            ws.cell(row_idx, c_price, price_int)
            if c_photos and photos:
                current_photos = str(ws.cell(row_idx, c_photos).value or "").strip()
                if current_photos != photos and _should_replace_photo_urls(
                    current_photos, source=resolved_photos.source
                ):
                    ws.cell(row_idx, c_photos, photos)
            if c_desc:
                ws.cell(
                    row_idx,
                    c_desc,
                    _format_description(
                        cfg.description_html,
                        nomenclature=nom,
                        article=article,
                        price=price_int,
                        quantity=stock_qty,
                        model_description=model_desc,
                        store_pitch=cfg.store_pitch_html,
                        store_defaults=row_defaults,
                        ushk_in_stock=_posting_ushk_in_stock(post),
                        sam_mb_cash_price=_posting_sam_mb_cash_price(post),
                    ),
                )
            if c_qty:
                ws.cell(row_idx, c_qty, AUTOLOAD_PRICE_QUANTITY)
            if c_brand and fields["brand"]:
                ws.cell(row_idx, c_brand, fields["brand"])
            if c_model and fields["model"]:
                ws.cell(row_idx, c_model, fields["model"])
            if c_w and fields["width"]:
                ws.cell(row_idx, c_w, fields["width"])
            if c_p and fields["profile"]:
                ws.cell(row_idx, c_p, fields["profile"])
            if c_d and fields["diameter"]:
                ws.cell(row_idx, c_d, fields["diameter"])
            if c_li and fields["load_index"]:
                ws.cell(row_idx, c_li, fields["load_index"])
            if c_si and fields["speed_index"]:
                ws.cell(row_idx, c_si, fields["speed_index"])
            if c_season and fields["season"]:
                ws.cell(row_idx, c_season, fields["season"])
            if c_run:
                ws.cell(row_idx, c_run, row_defaults.get("run_flat", "Нет"))
            if c_state:
                ws.cell(row_idx, c_state, row_defaults.get("condition", "Новое"))
            _apply_defaults(
                ws,
                row_idx,
                headers,
                row_defaults,
                force_keys=_STORE_FORCE_KEYS | frozenset({"merge_ads"}),
            )

    rounded = _apply_prices_to_sheet(
        ws,
        price_col=c_price,
        title_col=c_title,
        posting_df=posting_df,
    )
    if rounded:
        LOG.info("Цены в файле: обновлено/округлено %s строк", rounded)

    qty_updated = _apply_quantities_to_sheet(
        ws,
        qty_col=c_qty,
        title_col=c_title,
        posting_df=posting_df,
        max_quantity=cfg.max_listing_quantity,
    )
    if qty_updated:
        LOG.info(
            "Колонка «Количество»: исправлено %s строк → %s (цена за 1 шт)",
            qty_updated,
            AUTOLOAD_PRICE_QUANTITY,
        )

    photo_set, photo_cleared = _sync_photo_urls_to_sheet(
        ws,
        photos_col=c_photos,
        title_col=c_title,
        id_col=c_id,
        posting_df=posting_df,
        local_photos=local_photos,
        cfg=cfg,
        stores=stores,
        yandex_downloader=yandex_downloader,
    )
    if photo_set or photo_cleared:
        LOG.info(
            "Фото на Диске: проставлено %s, очищено чужих/пустых %s",
            photo_set,
            photo_cleared,
        )

    avito_set, avito_cleared = _sync_avito_ids_to_sheet(
        ws,
        avito_col=c_avito,
        id_col=c_id,
        title_col=c_title,
        posting_df=posting_df,
        avito_ids=avito_ids,
    )
    if avito_set or avito_cleared:
        LOG.info(
            "Номера Avito: проставлено %s, очищено чужих %s",
            avito_set,
            avito_cleared,
        )

    pruned_no_photos = 0
    if cfg.skip_without_photos and not cfg.include_all_goods_in_autoload:
        pruned_no_photos = _remove_rows_without_photos(
            ws,
            photos_col=c_photos,
            title_col=c_title,
        )
        if pruned_no_photos:
            LOG.info(
                "Без фото убрано из файла автозагрузки: %s строк (осталось с фото)",
                pruned_no_photos,
            )

    descriptions = _apply_descriptions_to_sheet(
        ws,
        desc_col=c_desc,
        title_col=c_title,
        price_col=c_price,
        id_col=c_id,
        qty_col=c_qty,
        posting_df=posting_df,
        cfg=cfg,
        stores=stores,
        model_descriptions=model_descriptions,
    )
    if descriptions:
        LOG.info("Описания в файле: обновлено %s строк", descriptions)

    merge_ads_set = _sync_merge_ads_to_sheet(ws, headers)
    if merge_ads_set:
        LOG.info("Мультиобъявление (MultiItem): включено в %s строках", merge_ads_set)

    multi_set, multi_groups = _sync_multi_names_to_sheet(ws, headers, title_col=c_title)
    if multi_set:
        LOG.info(
            "MultiName по размеру: обновлено %s строк, групп размеров: %s",
            multi_set,
            multi_groups,
        )

    saved_path = save_workbook(wb, output_path)
    stats["output_path"] = str(saved_path)
    stats["removed_rows"] = removed
    stats["photos_missing"] = photos_missing
    stats["missing_models"] = list(missing_models.values())
    stats["photos_dir"] = str(local_photos) if local_photos else ""
    return stats


def _row_has_avito_id(
    ws,
    row_idx: int,
    *,
    c_id: int | None,
    c_avito: int | None,
    avito_ids: dict[str, str],
) -> bool:
    avito_cell = ""
    if c_avito:
        avito_cell = str(ws.cell(row_idx, c_avito).value or "").strip()
        if avito_cell.endswith(".0"):
            avito_cell = avito_cell.split(".")[0]
        if avito_cell and avito_cell.lower() != "nan":
            return True
    listing_id = normalize_article_id(ws.cell(row_idx, c_id).value if c_id else None)
    article = _article_from_listing_id(listing_id) if listing_id else ""
    return bool(_avito_id_for_row(listing_id, article, avito_ids))


def filter_new_listings_workbook(
    source_path: Path,
    output_path: Path,
    *,
    avito_ids: dict[str, str],
) -> tuple[int, int]:
    """
    Фид для публикации: только новые объявления (без номера на Avito).

    Уже размещённые строки убираются — их цену/остаток обновляем через API.
    """
    wb = load_workbook(source_path)
    sheet_name = _find_data_sheet(wb)
    ws = wb[sheet_name]
    headers = _header_map(ws)
    c_id = _col(headers, "Уникальный идентификатор объявления")
    c_avito = _col(headers, "Номер объявления на Авито")
    c_title = _col(headers, "Название объявления")

    to_delete: list[int] = []
    kept = 0
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        title = str(ws.cell(row_idx, c_title).value or "").strip() if c_title else ""
        listing_id = normalize_article_id(ws.cell(row_idx, c_id).value if c_id else None)
        if not title and not listing_id:
            continue
        if _row_has_avito_id(ws, row_idx, c_id=c_id, c_avito=c_avito, avito_ids=avito_ids):
            to_delete.append(row_idx)
        else:
            kept += 1

    for row_idx in sorted(to_delete, reverse=True):
        ws.delete_rows(row_idx, 1)

    saved = save_workbook(wb, output_path)
    LOG.info(
        "Фид только новые: %s строк, убрано уже на Avito: %s → %s",
        kept,
        len(to_delete),
        saved.name,
    )
    return kept, len(to_delete)


def filter_photo_updates_workbook(
    source_path: Path,
    output_path: Path,
    *,
    avito_ids: dict[str, str],
) -> tuple[int, int]:
    """
    Уже на Avito + фото артикула в файле → отдельный фид для обновления (без дублей).

    Строка должна иметь AvitoId и URL с артикулом (не фото модели / CDN Авито).
    """
    wb = load_workbook(source_path)
    sheet_name = _find_data_sheet(wb)
    ws = wb[sheet_name]
    headers = _header_map(ws)
    c_id = _col(headers, "Уникальный идентификатор объявления")
    c_avito = _col(headers, "Номер объявления на Авито")
    c_title = _col(headers, "Название объявления")
    c_photos = _col(headers, "Ссылки на фото")

    to_delete: list[int] = []
    kept = 0
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        title = str(ws.cell(row_idx, c_title).value or "").strip() if c_title else ""
        listing_id = normalize_article_id(ws.cell(row_idx, c_id).value if c_id else None)
        if not title and not listing_id:
            continue
        if not _row_has_avito_id(ws, row_idx, c_id=c_id, c_avito=c_avito, avito_ids=avito_ids):
            to_delete.append(row_idx)
            continue
        article = _article_from_listing_id(listing_id) if listing_id else ""
        photos = str(ws.cell(row_idx, c_photos).value or "").strip() if c_photos else ""
        if not photo_urls_look_like_article(photos, article):
            to_delete.append(row_idx)
            continue
        kept += 1

    for row_idx in sorted(to_delete, reverse=True):
        ws.delete_rows(row_idx, 1)

    saved = save_workbook(wb, output_path)
    LOG.info(
        "Фид обновления фото: %s строк (уже на Avito + фото артикула) → %s",
        kept,
        saved.name,
    )
    return kept, len(to_delete)


def merge_autoload_feed_workbooks(
    sources: list[Path],
    output_path: Path,
) -> int:
    """Склеить несколько xlsx одного шаблона в один фид (строки данных подряд)."""
    paths = [p for p in sources if p.is_file()]
    if not paths:
        return 0
    shutil.copy2(paths[0], output_path)
    if len(paths) == 1:
        wb = load_workbook(output_path, read_only=True, data_only=True)
        try:
            ws = wb[_find_data_sheet(wb)]
            headers = _header_map(ws)
            c_title = _col(headers, "Название объявления")
            c_id = _col(headers, "Уникальный идентификатор объявления")
            count = 0
            for row_idx in range(DATA_START_ROW, ws.max_row + 1):
                title = str(ws.cell(row_idx, c_title).value or "").strip() if c_title else ""
                lid = normalize_article_id(ws.cell(row_idx, c_id).value if c_id else None)
                if title or lid:
                    count += 1
            return count
        finally:
            wb.close()

    wb = load_workbook(output_path)
    sheet_name = _find_data_sheet(wb)
    ws = wb[sheet_name]
    headers = _header_map(ws)
    c_title = _col(headers, "Название объявления")
    c_id = _col(headers, "Уникальный идентификатор объявления")
    max_col = ws.max_column
    last = DATA_START_ROW - 1
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        title = str(ws.cell(row_idx, c_title).value or "").strip() if c_title else ""
        lid = normalize_article_id(ws.cell(row_idx, c_id).value if c_id else None)
        if title or lid:
            last = row_idx
    next_row = last + 1
    total = last - DATA_START_ROW + 1

    for path in paths[1:]:
        wb2 = load_workbook(path, read_only=True, data_only=True)
        try:
            ws2 = wb2[_find_data_sheet(wb2)]
            h2 = _header_map(ws2)
            c_title2 = _col(h2, "Название объявления")
            c_id2 = _col(h2, "Уникальный идентификатор объявления")
            for row_idx in range(DATA_START_ROW, ws2.max_row + 1):
                title = str(ws2.cell(row_idx, c_title2).value or "").strip() if c_title2 else ""
                lid = normalize_article_id(ws2.cell(row_idx, c_id2).value if c_id2 else None)
                if not title and not lid:
                    continue
                for col in range(1, max_col + 1):
                    ws.cell(next_row, col, ws2.cell(row_idx, col).value)
                next_row += 1
                total += 1
        finally:
            wb2.close()

    save_workbook(wb, output_path)
    return total


_DEFAULT_HEADER_ALIASES = {
    "listing_type": "Способ размещения",
    "contact_person": "Контактное лицо",
    "phone": "Номер телефона",
    "address": "Адрес",
    "contact_method": "Способ связи",
    "category": "Категория",
    "goods_type": "Вид товара",
    "ad_type": "Вид объявления",
    "product_type": "Тип товара",
    "merge_ads": "Соединять это объявление с другими объявлениями",
    # альтернативные заголовки в новых шаблонах Авито — см. _MERGE_ADS_HEADERS
    "free_mounting": "Бесплатный шиномонтаж",
    "company": "Название компании",
    "email": "Почта",
    "audience": "Целевая аудитория",
}
